### Ingestion of the input data

## Import libraries
import csv
# import re
import numpy as np
import datetime as dt
import openpyxl as opxl
import code

## Settings
srcFile = 'raw_data/20170405221001.11876.events.csv'
startDate = dt.date(2009,1,1)
endDate = dt.date(2016,1,1)
totalDays = (endDate-startDate).days+1
dstFile = 'features/clean_20170405221001_11876.xlsx'

## Code
with open(srcFile, mode='r') as infile:
    reader = csv.reader(infile, delimiter=',')
    reader.next()
    # optionally do something with the header row

    dailyFrequencies = np.zeros([totalDays])
    for row in reader:
        date = dt.date(int(row[1][0:4]), int(row[1][4:6]), int(row[1][6:8]))
        daysSince = (date-startDate).days
        # print(daysSince)
        dailyFrequencies[daysSince] += 1


wb = opxl.Workbook()
ws = wb.active
ws.cell(row=1,column=1, value='Year')
ws.cell(row=1,column=2, value='Month')
ws.cell(row=1,column=3, value='Day')
ws.cell(row=1,column=4, value='Frequency')
for i in range(totalDays):
    date = (startDate+dt.timedelta(days=i))
    ws.cell(row=i+2,column=1, value=date.year)
    ws.cell(row=i+2,column=2, value=date.month)
    ws.cell(row=i+2,column=3, value=date.day)
    ws.cell(row=i+2,column=4, value=dailyFrequencies[i])
wb.save(dstFile)
# code.interact(local=locals())
# exit()
